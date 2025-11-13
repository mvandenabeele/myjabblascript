import os
import myjabbla
from dotenv import load_dotenv
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from threading import Lock

def select_subgroup(parent_group: myjabbla.Group) -> myjabbla.Group:
    target_subgroups = parent_group.subgroups()
    if len(target_subgroups) > 0:
        print("The target group has subgroups, please select one:")
        print("0: <Select this group>")
        for idx, sg in enumerate(target_subgroups):
            print(f"{idx+1}: {sg.name} (id:{sg.id})")
        sel = int(input("Select subgroup by number: "))
        if sel == 0:
            return None
        if sel < 1 or sel > len(target_subgroups):
            print("Invalid selection")
            return

        subgroup = select_subgroup(target_subgroups[sel-1])
        if subgroup is not None:
            return subgroup
        else:
            return target_subgroups[sel-1]

    return None

def check_user_exists(server, login, row_index):
    """Helper function to check if a user exists (for multithreading)"""
    try:
        exists_user = server.get_user(login)
        return row_index, exists_user, None
    except myjabbla.ItemNotFoundError:
        return row_index, None, None
    except myjabbla.ApiError as e:
        return row_index, None, e

def create_user_account(target_group, login, password, email, row_index):
    """Helper function to create a user account (for multithreading)"""
    try:
        added_user = target_group.add_user(login, password, email)
        return row_index, added_user, None
    except myjabbla.ApiError as e:
        return row_index, None, e

def load_lines_from_xlsx(file_path):
    from openpyxl import load_workbook
    wb = load_workbook(filename=file_path)
    sheet = wb.active
    
    max_line_width = 0
    lines = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        line_width = 0
        for w, cell in enumerate(row):
            if cell is not None:
                line_width = w
                
        if line_width > max_line_width:
            max_line_width = line_width
                    
        has_data = line_width > 0

        if has_data:
            lines.append(row)
    
    clean_lines = []
    for row in lines:
        clean_lines.append( row[:max_line_width+1] )
    return clean_lines

def load_lines_from_csv(file_path):
    import csv
    lines = []
    with open(file_path) as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            row = tuple(row)
            lines.append(row)
    return lines

def process_xlsx(file_path, target_group: myjabbla.Group, server: myjabbla.Server, max_workers=10):
    
    if file_path.endswith(".xlsx"):
        data_lines = load_lines_from_xlsx(file_path)
    else:
        data_lines = load_lines_from_csv(file_path)

    # print first 10 rows
    for i, row in enumerate(data_lines):
        if i < 10:
            print(i, row)
        else:
            break
        
    #ask for line number to start from
    start_line = int(input("Enter line number to start from (0-indexed): "))
    data_lines = data_lines[start_line:]
    
    login_col = None
    pwd_col = None
    email_col = None
    
    for i, row in enumerate(data_lines):
        print(f"----- record {i} -----")
        for j,c in enumerate(row):
            print(j, c)

    login_col = int(input("Enter column number for login: "))
    pwd_col = int(input("Enter column number for password: "))
    email_col = int(input("Enter column number for email (or -1 if none): "))   
    
    # Collect all user data first
    user_data = []
    for i, row in enumerate(data_lines):
        if row[login_col] and row[pwd_col] and row[email_col]:  # Skip empty rows
            user_data.append((i+start_line, row[login_col], row[pwd_col], row[email_col] if email_col >= 0 else ""))
    
    print(f"Found {len(user_data)} users to process")
    
    # Multithreaded user existence checking
    conflicts = []
    errors = []
    lock = Lock()
    
    print("Checking existing users...")
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all check tasks
        future_to_data = {
            executor.submit(check_user_exists, server, login, row_index): (row_index, login)
            for row_index, login, _, _ in user_data
        }
        
        completed = 0
        for future in as_completed(future_to_data):
            row_index, existing_user, error = future.result()
            completed += 1
            
            if completed % 10 == 0 or completed == len(user_data):
                print(f"Checked {completed}/{len(user_data)} users...")
            
            if existing_user:
                with lock:
                    conflicts.append(existing_user)
            elif error:
                with lock:
                    errors.append((row_index, future_to_data[future][1], error.message))
    
    if errors:
        print("Errors occurred while checking users:")
        for row_index, login, error_msg in errors:
            print(f" - Row {row_index}, login {login}: {error_msg}")
    
    if len(conflicts) > 0:
        print("The following logins already exist:")
        for c in conflicts:
            print(f" - {c.login} (id:{c.id}, packet:{c.packet_sn})")
            
        print("\nPlease resolve these conflicts and try again.")
        return
    
    print("No conflicts found, proceeding with import.")
    ok = input("Type 'yes' to proceed: ")
    if ok.lower() != 'yes':
        print("Aborting.")
        return
    
    # Multithreaded user creation
    creation_errors = []
    created_users = []
    
    print("Creating users...")
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all creation tasks
        future_to_data = {
            executor.submit(create_user_account, target_group, login, password, email, row_index): 
            (row_index, login, password, email)
            for row_index, login, password, email in user_data
        }
        
        completed = 0
        for future in as_completed(future_to_data):
            row_index, created_user, error = future.result()
            completed += 1
            
            if completed % 10 == 0 or completed == len(user_data):
                print(f"Created {completed}/{len(user_data)} users...")
            
            if created_user:
                with lock:
                    created_users.append((row_index, created_user))
            else:
                with lock:
                    creation_errors.append((row_index, future_to_data[future][1], error.message))
    
    # Report results
    print(f"\nImport completed!")
    print(f"Successfully created: {len(created_users)} users")
    
    if creation_errors:
        print(f"Errors creating {len(creation_errors)} users:")
        for row_index, login, error_msg in creation_errors:
            print(f" - Row {row_index}, login {login}: {error_msg}")
              
def main():
    load_dotenv()
    
    # Configuration for multithreading
    MAX_WORKERS = 10  # Adjust this based on your API rate limits
    

    # https://patorjk.com/software/taag/#p=display&f=Ogre&t=MyJabbla+Bulk+Impoerter&x=none&v=4&h=4&w=80&we=false
    print(
'''                                       
               __        _     _     _           ___       _ _       _____                            _            
  /\/\  _   _  \ \  __ _| |__ | |__ | | __ _    / __\_   _| | | __   \_   \_ __ ___  _ __   ___  _ __| |_ ___ _ __ 
 /    \| | | |  \ \/ _` | '_ \| '_ \| |/ _` |  /__\// | | | | |/ /    / /\/ '_ ` _ \| '_ \ / _ \| '__| __/ _ \ '__|
/ /\/\ \ |_| /\_/ / (_| | |_) | |_) | | (_| | / \/  \ |_| | |   <  /\/ /_ | | | | | | |_) | (_) | |  | ||  __/ |   
\/    \/\__, \___/ \__,_|_.__/|_.__/|_|\__,_| \_____/\__,_|_|_|\_\ \____/ |_| |_| |_| .__/ \___/|_|   \__\___|_|   
        |___/                                                                       |_|                            
'''
    )
    print(f"Using base url {os.getenv('MYJABBLA_BASE_URL')}")
    
    mj = myjabbla.Server(os.getenv("MYJABBLA_BASE_URL"))
    mj.set_api_key(os.getenv("MYJABBLA_API_KEY"))
    
    file_candidates = [] 
    # find first xlsx file in current directory
    for file in os.listdir("."):
        if file.endswith(".csv"):
            print(f"Found csv file: {file}")
            file_candidates.append(file)
        elif file.endswith(".xlsx"):
            print(f"Found xlsx file: {file}")
            file_candidates.append(file)
            
    for idx, fname in enumerate(file_candidates):
        print(f"{idx}: {fname}")
    file = None
    if len(file_candidates) == 0:
        print("No xlsx or csv files found in current directory. Please place the file to import here.")
        return
    elif len(file_candidates) == 1:
        file = file_candidates[0]
    else:
        sel = int(input("Multiple files found, select file by number: "))
        if sel < 0 or sel >= len(file_candidates):
            print("Invalid selection")
            return
        file = file_candidates[sel]
        
    if file:
        print(f"Processing file: {file}")
        
    packet = input("What serial number should accounts be added to? ")
    try:
        target_group = mj.get_group_sn(packet)
        target_group = select_subgroup(target_group) or target_group
        
        print(target_group)
        process_xlsx(file, target_group, mj, MAX_WORKERS)
            
    except myjabbla.ApiError as e:  
        print(f"Error: {e.message}")    
        

if __name__ == "__main__":
    main()